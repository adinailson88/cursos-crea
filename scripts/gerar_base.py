"""
gerar_base.py

Le os exports brutos em dados/fonte/*.csv (extraidos do Google Drive do CREA-BA)
e gera a planilha unica normalizada planilha/Base_IES_Cursos_CREA_BA.xlsx com as
abas: IES, CAMPI, CURSOS, TITULOS_PROFISSIONAIS, TITULOS_EQUIVALENCIAS,
CURSO_TITULO, BASE_LEGAL, ATRIBUICOES, TITULO_ATRIBUICAO_LEGAL, FONTES,
PENDENCIAS, BASE_DASHBOARD.

Principios aplicados (ver README.md e o prompt original):
 - nada e inventado: o que nao pode ser confirmado na fonte recebe
   "PENDENTE DE VALIDACAO";
 - nenhum registro e apagado silenciosamente; duplicados e casos
   historicos/transferidos/desativados sao preservados com situacao marcada;
 - identificadores (CNPJ, codigos) sao sempre tratados como texto.

Uso:
    python scripts/gerar_base.py
"""
import csv
import hashlib
import os
import re
import unicodedata
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FONTE_DIR = os.path.join(BASE_DIR, "dados", "fonte")
PLANILHA_DIR = os.path.join(BASE_DIR, "planilha")
OUT_XLSX = os.path.join(PLANILHA_DIR, "Base_IES_Cursos_CREA_BA.xlsx")

PENDENTE = "PENDENTE DE VALIDAÇÃO"
NAO_INFORMADA = "NÃO INFORMADA"
RUN_TIMESTAMP = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
RUN_DATE = datetime.now().strftime("%Y-%m-%d")

# ---------------------------------------------------------------------------
# Utilitarios de normalizacao
# ---------------------------------------------------------------------------

def norm_ws(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s))
    s = s.replace("﻿", "")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def strip_accents_upper(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.upper().strip()


def cnpj_digits(s):
    return re.sub(r"\D", "", s or "")


def cnpj_dv_valid(cnpj):
    c = cnpj_digits(cnpj)
    if len(c) != 14:
        return False
    if c == c[0] * 14:
        return False

    def calc(base, weights):
        total = sum(int(d) * w for d, w in zip(base, weights))
        r = total % 11
        return "0" if r < 2 else str(11 - r)

    w1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    d1 = calc(c[:12], w1)
    w2 = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    d2 = calc(c[:12] + d1, w2)
    return c[-2:] == d1 + d2


def stable_id(prefix, *parts, n=8):
    h = hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:n].upper()
    return f"{prefix}-{h}"


CODIGO_PREFIXO_RE = re.compile(r"^\s*(\d{1,10})\s*-\s*(.+)$")


def extrai_codigo_e_nome(nome_raw):
    """Separa 'CODIGO - NOME' quando presente no inicio do nome de origem."""
    nome_raw = norm_ws(nome_raw)
    m = CODIGO_PREFIXO_RE.match(nome_raw)
    if m:
        return m.group(1), norm_ws(m.group(2))
    return "", nome_raw


SIGLA_RE = re.compile(r"-\s*([A-ZÀ-Úa-zà-ú0-9\.]{2,20})\s*$")


def extrai_sigla(nome_normalizado):
    m = SIGLA_RE.search(nome_normalizado)
    if not m:
        return ""
    candidato = m.group(1).strip(" .")
    letras = re.sub(r"[^A-Za-zÀ-Úà-ú]", "", candidato)
    if len(letras) < 2:
        return ""
    if candidato.upper() != candidato and not any(ch.isupper() for ch in candidato):
        return ""
    return candidato.upper()


MARCADORES_SITUACAO = [
    (re.compile(r"\[?\s*TRANSFERIDO[A-ZÇÃÕÁÉÍÓÚ ]*OUTRO CONSELHO\s*\]?", re.IGNORECASE), "TRANSFERIDO"),
    (re.compile(r"\[?\s*NOME\s+ANTIGO\s*[:;]?\s*.*$", re.IGNORECASE), "HISTÓRICO"),
    (re.compile(r"\bDESATIVAD[AO]\b", re.IGNORECASE), "INATIVO"),
    (re.compile(r"\[?\s*SEM IDENTIFICA[ÇC][ÃA]O\s*\]?", re.IGNORECASE), "NÃO IDENTIFICADO"),
    (re.compile(r"\[?\s*SEM CADASTRO\s*\]?", re.IGNORECASE), "NÃO IDENTIFICADO"),
    (re.compile(r"\[?\s*DESCREDENCIADA[A-ZÇÃÕÁÉÍÓÚ ]*\]?", re.IGNORECASE), "INATIVO"),
    (re.compile(r"CADASTRO SUSPENSO[A-ZÇÃÕÁÉÍÓÚ\- ]*", re.IGNORECASE), "INATIVO"),
    (re.compile(r"^\s*EXCLUIR\s*$", re.IGNORECASE), "NÃO IDENTIFICADO"),
]

# Correções pontuais de grafia/variação conhecidas do campus -> nome oficial IBGE
# (correções objetivas de digitação/alcunha, não deduções de local).
ALIASES_MUNICIPIO = {
    strip_accents_upper(k): strip_accents_upper(v) for k, v in {
        "TANCREDO NEVES": "PRESIDENTE TANCREDO NEVES",
        "CRUS DAS ALMAS": "CRUZ DAS ALMAS",
        "DIAS D AVILA": "DIAS D'ÁVILA",
        "DIAS D'AVILA": "DIAS D'ÁVILA",
        "IHEUS": "ILHÉUS",
        "EUCLIDES CUNHA": "EUCLIDES DA CUNHA",
        "XIQUE - XIQUE": "XIQUE-XIQUE",
        "XIQUE XIQUE": "XIQUE-XIQUE",
    }.items()
}


def detecta_situacao_e_limpa(campus_original):
    situacao = None
    campus_limpo = campus_original
    for regex, label in MARCADORES_SITUACAO:
        if regex.search(campus_limpo):
            situacao = label
            campus_limpo = regex.sub("", campus_limpo)
    campus_limpo = norm_ws(campus_limpo)
    if not campus_limpo and situacao is None:
        situacao = "NÃO IDENTIFICADO"
    if situacao is None:
        situacao = "SITUAÇÃO NÃO INFORMADA"
    return situacao, campus_limpo


def carrega_municipios_bahia():
    path = os.path.join(FONTE_DIR, "municipios_bahia_ibge.csv")
    lookup = {}
    with open(path, encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            nome = row["nome"]
            key = strip_accents_upper(nome)
            lookup[key] = {
                "nome": nome,
                "codigo_ibge": row["codigo_ibge"],
                "lat": float(row["latitude"]),
                "lon": float(row["longitude"]),
            }
    return lookup


SUFIXOS_CAMPUS = re.compile(
    r"\s*-\s*(BA|CAMPUS[:\s].*|PARALELA|STIEP|BARBALHO|CANDEIAS|PITUBA/MERC[ÊE]S)\s*$",
    re.IGNORECASE,
)


def identifica_municipio(campus_limpo, municipios_lookup):
    """Tenta casar o texto de campus com um municipio oficial da Bahia.
    Retorna (municipio_nome_ibge, info) ou (None, None) quando nao confirmado.
    """
    if not campus_limpo:
        return None, None
    texto = campus_limpo
    texto = re.sub(r"CAMPUS\s*:?.*$", "", texto, flags=re.IGNORECASE).strip()
    texto = re.sub(r"\(CAMPUS[^)]*\)", "", texto, flags=re.IGNORECASE).strip()
    candidatos = re.split(r"\s*[-/]\s*", texto)
    candidatos = [norm_ws(c) for c in candidatos if norm_ws(c)]
    # tenta o texto inteiro primeiro, depois cada token da esquerda para a direita
    tentativas = [texto] + candidatos
    for cand in tentativas:
        key = strip_accents_upper(cand)
        if key in ALIASES_MUNICIPIO:
            key = ALIASES_MUNICIPIO[key]
        if key in municipios_lookup:
            return municipios_lookup[key]["nome"], municipios_lookup[key]
    # tenta remover sufixos comuns de bairro/campus e tentar de novo so com o 1o token
    if candidatos:
        key = strip_accents_upper(candidatos[0])
        if key in ALIASES_MUNICIPIO:
            key = ALIASES_MUNICIPIO[key]
        if key in municipios_lookup:
            return municipios_lookup[key]["nome"], municipios_lookup[key]
    return None, None


CATEGORIAS_ORIGEM = {"PÚBLICA": "PÚBLICA", "PRIVADA": "PRIVADA"}


def normaliza_categoria(cat_original):
    c = strip_accents_upper(cat_original)
    if c in ("PUBLICA",):
        return "PÚBLICA"
    if c in ("PRIVADA",):
        return "PRIVADA"
    return NAO_INFORMADA


def infere_natureza_administrativa(nome_normalizado, categoria):
    nome_up = strip_accents_upper(nome_normalizado)
    if categoria == "PÚBLICA":
        if "FEDERAL" in nome_up:
            return "PÚBLICA FEDERAL"
        if "ESTADUAL" in nome_up:
            return "PÚBLICA ESTADUAL"
        if "MUNICIPAL" in nome_up:
            return "PÚBLICA MUNICIPAL"
        return PENDENTE
    if categoria == "PRIVADA":
        return "PRIVADA"
    return PENDENTE


# ---------------------------------------------------------------------------
# Estruturas de saida
# ---------------------------------------------------------------------------
IES_ROWS = []          # dict por id_ies
CAMPI_ROWS = []        # dict por id_campus
CURSOS_ROWS = []
TITULOS_ROWS = []
TITULOS_EQUIV_ROWS = []
CURSO_TITULO_ROWS = []
BASE_LEGAL_ROWS = []
ATRIBUICOES_ROWS = []
TITULO_ATRIB_LEGAL_ROWS = []
FONTES_ROWS = []
PENDENCIAS_ROWS = []

_pendencia_seq = 0


def add_pendencia(tipo_registro, id_registro, campo, problema, gravidade, acao, observacao=""):
    global _pendencia_seq
    _pendencia_seq += 1
    PENDENCIAS_ROWS.append({
        "id_pendencia": f"PEND-{_pendencia_seq:05d}",
        "tipo_registro": tipo_registro,
        "id_registro": id_registro,
        "campo": campo,
        "problema": problema,
        "gravidade": gravidade,
        "acao_recomendada": acao,
        "status": "ABERTA",
        "observacao": observacao,
    })


# ---------------------------------------------------------------------------
# 1) FONTES
# ---------------------------------------------------------------------------

def registra_fonte(id_fonte, nome_arquivo, tipo_fonte, data_emissao, descricao, prioridade):
    FONTES_ROWS.append({
        "id_fonte": id_fonte,
        "nome_arquivo": nome_arquivo,
        "tipo_fonte": tipo_fonte,
        "data_emissao": data_emissao,
        "data_importacao": RUN_DATE,
        "descricao": descricao,
        "observacao": "",
        "prioridade_fonte": prioridade,
    })


registra_fonte("FONTE-01", "Relatorio_Instituicoes.csv (Relatório Instituições)", "RELATÓRIO SITAC",
                "2025-08-06", "Relatório genérico SITAC com instituições/campi cadastrados no CREA-BA. Fonte institucional principal.", 1)
registra_fonte("FONTE-02", "Camara_Eletrica.csv (Relação de Instituições da Câmara de Elétrica)", "RELATÓRIO SITAC",
                "2025-08-06", "Subconjunto de instituições já vinculadas à Câmara Especializada de Engenharia Elétrica. Uso: cruzamento/validação.", 2)
registra_fonte("FONTE-03", "Camara_Eletrica_modalidade.csv (Relação de Instituições da Câmara de Elétrica por modalidade)", "RELATÓRIO SITAC",
                "2025-08-06", "Mesmo subconjunto de FONTE-02, exportado em outro momento. Uso: cruzamento/validação.", 2)
registra_fonte("FONTE-04", "Titulos_SITAC.csv (sitac_relatorios_relatorio_generico_08-08-2025)", "RELATÓRIO SITAC",
                "2025-08-08", "Catálogo de títulos profissionais (código, tratamento, nível) usado no SITAC do CREA-BA.", 1)
registra_fonte("FONTE-05", "Consolidado_EXEMPLO.csv (aba EXEMPLO do arquivo Consolidado)", "PLANILHA DE TRABALHO",
                "2025-09-09", "7 registros de exemplo curso+instituição+título+endereço, fornecidos como modelo de estrutura pretendida. Não representam cadastro validado.", 3)
registra_fonte("FONTE-06", "municipios_bahia_ibge.csv (kelvins/municipios-brasileiros, dados derivados do IBGE)", "REFERÊNCIA GEOGRÁFICA EXTERNA",
                RUN_DATE, "Lista dos 417 municípios da Bahia com centroide (latitude/longitude), usada para geocodificação municipal quando o campus não tem coordenada própria.", 2)
registra_fonte("FONTE-07", "Lei nº 5.194/1966 (planalto.gov.br)", "LEGISLAÇÃO FEDERAL", "1966-12-24",
                "Regula o exercício das profissões de Engenheiro, Arquiteto e Engenheiro-Agrônomo. Base legal geral das atribuições.", 1)
registra_fonte("FONTE-08", "Resolução CONFEA nº 218/1973 (normativos.confea.org.br)", "NORMATIVO CONFEA", "1973-06-29",
                "Discrimina atividades das diferentes modalidades profissionais da Engenharia, Arquitetura e Agronomia. Arts. 8º e 9º revogados pela Resolução 1.156/2025.", 1)
registra_fonte("FONTE-09", "Resolução CONFEA nº 1.156/2025 (normativos.confea.org.br)", "NORMATIVO CONFEA", "2025-10-24",
                "Consolida normativos das atividades e competências profissionais dos engenheiros da modalidade eletricista (eletricista, eletrônica, computação, automação, energia, software, biomédica, industrial elétrica).", 1)

print("FONTES: %d registradas" % len(FONTES_ROWS))

# ---------------------------------------------------------------------------
# 2) IES + CAMPI  (fonte principal: Relatorio_Instituicoes.csv)
# ---------------------------------------------------------------------------
MUNICIPIOS_BAHIA = carrega_municipios_bahia()

def le_linhas_relatorio(caminho):
    with open(caminho, encoding="utf-8", newline="") as f:
        reader = list(csv.reader(f))
    hdr_idx = next(i for i, r in enumerate(reader) if len(r) >= 2 and r[0].strip() == "NOME" and r[1].strip() == "CAMPUS")
    linhas = reader[hdr_idx + 1:]
    linhas = [r for r in linhas if any((c or "").strip() for c in r)]
    return linhas


linhas_inst = le_linhas_relatorio(os.path.join(FONTE_DIR, "Relatorio_Instituicoes.csv"))
print("Relatorio_Instituicoes.csv: %d linhas de dados" % len(linhas_inst))

# ies_key -> id_ies ; agrupamento por CNPJ valido (nao ficticio) quando presente,
# senao por nome_normalizado (ver README: CNPJ da fonte se repete entre campi da
# mesma instituicao, entao serve para agrupar IES; nome isolado NAO e usado como
# chave de campus, apenas de fallback quando nao ha CNPJ).
ies_by_key = {}
ies_linha_contagem = {}
campus_dedup = {}  # (id_ies, campus_normalizado) -> id_campus  (deduplicacao de linhas repetidas)
linha_origem_counter = 0

for row in linhas_inst:
    linha_origem_counter += 1
    row = (row + [""] * 6)[:6]
    nome_raw, campus_raw, cnpj_raw, codigo_mec_raw, codigo_regional_raw, categoria_raw = row

    nome_raw = norm_ws(nome_raw)
    if not nome_raw:
        continue

    codigo_do_nome, nome_sem_codigo = extrai_codigo_e_nome(nome_raw)
    nome_normalizado = norm_ws(re.sub(r"\s{2,}", " ", nome_sem_codigo))
    sigla = extrai_sigla(nome_normalizado)

    cnpj_norm = cnpj_digits(cnpj_raw)
    cnpj_ficticio = cnpj_norm == "00000000000000"
    cnpj_valido = (not cnpj_ficticio) and cnpj_dv_valid(cnpj_norm) if cnpj_norm else False

    codigo_mec = norm_ws(codigo_mec_raw) or codigo_do_nome

    categoria = normaliza_categoria(categoria_raw)

    # chave de agrupamento da IES
    if cnpj_norm and not cnpj_ficticio and len(cnpj_norm) == 14:
        ies_key = "CNPJ:" + cnpj_norm
    else:
        ies_key = "NOME:" + strip_accents_upper(nome_normalizado)

    if ies_key not in ies_by_key:
        id_ies = stable_id("IES", ies_key)
        natureza = infere_natureza_administrativa(nome_normalizado, categoria)
        ies_by_key[ies_key] = id_ies
        IES_ROWS.append({
            "id_ies": id_ies,
            "nome_original": nome_raw,
            "nome_normalizado": nome_normalizado,
            "sigla": sigla,
            "cnpj_original": cnpj_raw,
            "cnpj_normalizado": cnpj_norm,
            "cnpj_valido": "SIM" if cnpj_valido else ("FICTÍCIO" if cnpj_ficticio else ("NÃO" if cnpj_norm else PENDENTE)),
            "codigo_mec": codigo_mec or PENDENTE,
            "categoria_academica": categoria,
            "natureza_administrativa": natureza,
            "situacao_ies": "SITUAÇÃO NÃO INFORMADA",
            "observacao_cadastral": "",
            "fonte_origem": "FONTE-01",
            "data_fonte": "2025-08-06",
            "publicar_dashboard": "SIM",
        })
        ies_linha_contagem[ies_key] = 0
        if not codigo_mec:
            add_pendencia("IES", id_ies, "codigo_mec", "Instituição sem Código MEC", "MÉDIA",
                          "Confirmar Código MEC no cadastro SITAC ou portal e-MEC.")
        if not cnpj_norm:
            add_pendencia("IES", id_ies, "cnpj_normalizado", "Instituição sem CNPJ", "MÉDIA",
                          "Confirmar CNPJ no cadastro SITAC.")
        elif cnpj_ficticio:
            add_pendencia("IES", id_ies, "cnpj_normalizado", "CNPJ fictício (00000000000000)", "ALTA",
                          "Localizar CNPJ real da instituição antes de publicar.")
        elif not cnpj_valido:
            add_pendencia("IES", id_ies, "cnpj_normalizado", "CNPJ com dígito verificador inválido", "ALTA",
                          "Revisar digitação do CNPJ na fonte original.")
        if categoria == NAO_INFORMADA:
            add_pendencia("IES", id_ies, "categoria_academica", "Categoria acadêmica não informada na fonte", "BAIXA",
                          "Confirmar categoria (pública/privada) no cadastro SITAC.")
    else:
        id_ies = ies_by_key[ies_key]
        # segunda ocorrencia do CNPJ com nome levemente diferente -> registra variação
        ies_row = next(r for r in IES_ROWS if r["id_ies"] == id_ies)
        if strip_accents_upper(nome_normalizado) != strip_accents_upper(ies_row["nome_normalizado"]):
            nota = f"Variação de nome observada na fonte: \"{nome_raw}\"."
            if nota not in ies_row["observacao_cadastral"]:
                ies_row["observacao_cadastral"] = norm_ws(ies_row["observacao_cadastral"] + " " + nota)

    ies_linha_contagem[ies_key] += 1

    # ---- CAMPUS ----
    campus_raw_ws = norm_ws(campus_raw)
    situacao_campus, campus_limpo = detecta_situacao_e_limpa(campus_raw_ws)
    campus_normalizado = campus_limpo or campus_raw_ws

    municipio_nome, municipio_info = (None, None)
    if situacao_campus != "TRANSFERIDO":
        municipio_nome, municipio_info = identifica_municipio(campus_normalizado, MUNICIPIOS_BAHIA)

    campus_key = (id_ies, strip_accents_upper(campus_normalizado), situacao_campus)
    if campus_key in campus_dedup:
        id_campus = campus_dedup[campus_key]
        camp_row = next(r for r in CAMPI_ROWS if r["id_campus"] == id_campus)
        camp_row["_repeticoes"] = camp_row.get("_repeticoes", 1) + 1
        continue

    id_campus = stable_id("CAMP", id_ies, str(len(campus_dedup)))
    campus_dedup[campus_key] = id_campus

    lat = lon = ""
    origem_coordenada = ""
    if municipio_info:
        lat = municipio_info["lat"]
        lon = municipio_info["lon"]
        origem_coordenada = "centroide do município (IBGE)"

    CAMPI_ROWS.append({
        "id_campus": id_campus,
        "id_ies": id_ies,
        "campus_original": campus_raw_ws,
        "campus_normalizado": campus_normalizado,
        "municipio": municipio_nome or "",
        "uf": "BA" if municipio_nome else "",
        "endereco": "",
        "bairro": "",
        "cep": "",
        "latitude": lat,
        "longitude": lon,
        "origem_coordenada": origem_coordenada,
        "situacao_campus": situacao_campus,
        "observacao": "",
        "publicar_dashboard": "SIM" if situacao_campus == "SITUAÇÃO NÃO INFORMADA" else "NÃO",
        "_repeticoes": 1,
        "_linha_origem": linha_origem_counter,
    })

    if situacao_campus not in ("TRANSFERIDO",) and not municipio_nome and campus_normalizado:
        add_pendencia("CAMPUS", id_campus, "municipio", "Não foi possível identificar o município do campus a partir do texto de origem", "MÉDIA",
                      "Revisar manualmente o texto do campus e confirmar o município.", observacao=f"Texto original: \"{campus_raw_ws}\"")
    if not campus_normalizado:
        add_pendencia("CAMPUS", id_campus, "campus_normalizado", "Campus sem identificação na fonte", "MÉDIA",
                      "Confirmar unidade/campus no cadastro SITAC.")

for row in CAMPI_ROWS:
    reps = row.pop("_repeticoes", 1)
    if reps > 1:
        nota = f"Linha repetida {reps}x na fonte original (mesma instituição/campus/situação)."
        row["observacao"] = norm_ws((row["observacao"] + " " + nota).strip())
        add_pendencia("CAMPUS", row["id_campus"], "linha_origem", "Registro duplicado na fonte", "BAIXA",
                      "Duplicidade preservada e consolidada em um único registro de campus; sem perda de dado (fonte bruta arquivada em dados/fonte/).",
                      observacao=f"{reps} ocorrências na fonte original.")
    row.pop("_linha_origem", None)

print("IES: %d instituições únicas | CAMPI: %d campi únicos (a partir de %d linhas de origem)" % (
    len(IES_ROWS), len(CAMPI_ROWS), len(linhas_inst)))

# ---------------------------------------------------------------------------
# 3) Cruzamento com as planilhas da Câmara de Elétrica (validação, não fonte primária)
# ---------------------------------------------------------------------------
CNPJ_TO_IES = {r["cnpj_normalizado"]: r for r in IES_ROWS if r["cnpj_normalizado"]}


def cruza_camara_eletrica(caminho, rotulo_fonte):
    linhas = le_linhas_relatorio(caminho)
    vinculadas = 0
    for row in linhas:
        row = (row + [""] * 6)[:6]
        nome_raw, campus_raw, cnpj_raw, codigo_mec_raw, _, categoria_raw = row
        cnpj_norm = cnpj_digits(cnpj_raw)
        if not cnpj_norm:
            continue
        ies_row = CNPJ_TO_IES.get(cnpj_norm)
        if not ies_row:
            continue
        vinculadas += 1
        nota = f"Vinculada à Câmara de Elétrica (fonte: {rotulo_fonte})."
        if nota not in ies_row["observacao_cadastral"]:
            ies_row["observacao_cadastral"] = norm_ws(ies_row["observacao_cadastral"] + " " + nota)
        categoria_camara = normaliza_categoria(categoria_raw)
        if categoria_camara != NAO_INFORMADA and categoria_camara != ies_row["categoria_academica"]:
            add_pendencia("IES", ies_row["id_ies"], "categoria_academica",
                          f"Divergência de categoria acadêmica entre fontes ({rotulo_fonte}: {categoria_camara} × Relatório Instituições: {ies_row['categoria_academica']})",
                          "MÉDIA", "Confirmar categoria correta no cadastro SITAC.")
    return vinculadas


v1 = cruza_camara_eletrica(os.path.join(FONTE_DIR, "Camara_Eletrica.csv"), "Relação de Instituições da Câmara de Elétrica")
v2 = cruza_camara_eletrica(os.path.join(FONTE_DIR, "Camara_Eletrica_modalidade.csv"), "Relação de Instituições da Câmara de Elétrica por modalidade")
print("Câmara Elétrica: %d + %d vínculos por CNPJ confirmados" % (v1, v2))

# ---------------------------------------------------------------------------
# 4) TITULOS_PROFISSIONAIS + TITULOS_EQUIVALENCIAS  (fonte: Titulos_SITAC.csv)
# ---------------------------------------------------------------------------

def le_titulos(caminho):
    with open(caminho, encoding="utf-8", newline="") as f:
        reader = list(csv.reader(f))
    hdr_idx = next(i for i, r in enumerate(reader) if len(r) >= 4 and r[0].strip() == "MODALIDADE" and r[1].strip() == "TÍTULO")
    linhas = reader[hdr_idx + 1:]
    linhas = [r for r in linhas if any((c or "").strip() for c in r)]
    return linhas


linhas_titulos = le_titulos(os.path.join(FONTE_DIR, "Titulos_SITAC.csv"))
print("Titulos_SITAC.csv: %d linhas de dados" % len(linhas_titulos))

por_codigo = {}
for row in linhas_titulos:
    row = (row + [""] * 5)[:5]
    modalidade, titulo, tratamento, codigo, nivel = [norm_ws(c) for c in row]
    codigo = re.sub(r"\D", "", codigo)
    if not codigo:
        continue
    por_codigo.setdefault(codigo, []).append({
        "modalidade": modalidade, "titulo": titulo, "tratamento": tratamento, "nivel": nivel,
    })


def divide_feminino_masculino(tratamento, masculino_ref):
    """Usa o titulo (coluna TÍTULO, forma masculina) para separar o TRATAMENTO
    (feminino+masculino concatenados sem separador) em duas strings."""
    if not tratamento:
        return "", ""
    if not masculino_ref:
        return "", tratamento
    if tratamento.endswith(masculino_ref):
        return norm_ws(tratamento[: -len(masculino_ref)]), masculino_ref
    # comparacao tolerante a acentos/caixa para achar o ponto de corte
    trat_up = strip_accents_upper(tratamento)
    masc_up = strip_accents_upper(masculino_ref)
    if trat_up.endswith(masc_up):
        corte = len(tratamento) - len(masculino_ref)
        return norm_ws(tratamento[:corte]), norm_ws(tratamento[corte:])
    return "", tratamento


for codigo, ocorrencias in sorted(por_codigo.items()):
    preenchidas = [o for o in ocorrencias if o["titulo"]]
    vazias = [o for o in ocorrencias if not o["titulo"]]
    if preenchidas:
        principal = preenchidas[0]
        extras = preenchidas[1:]
    else:
        principal = ocorrencias[0]
        extras = ocorrencias[1:]

    titulo_masc = principal["titulo"]
    fem, masc = divide_feminino_masculino(principal["tratamento"], titulo_masc)
    titulo_normalizado = titulo_masc or "NÃO INFORMADO"
    if fem and masc and strip_accents_upper(fem) != strip_accents_upper(masc):
        titulo_display = f"{masc} / {fem}"
    else:
        titulo_display = titulo_masc or "NÃO INFORMADO"

    TITULOS_ROWS.append({
        "codigo_titulo": codigo,
        "titulo_normalizado": titulo_normalizado,
        "titulo_feminino": fem,
        "titulo_masculino": masc or titulo_masc,
        "nivel_formacao": principal["nivel"] or PENDENTE,
        "modalidade_profissional": principal["modalidade"] or PENDENTE,
        "situacao_titulo": "ATIVO NO CATÁLOGO" if titulo_masc else PENDENTE,
        "fonte_origem": "FONTE-04",
        "observacao": "Título ausente no catálogo de origem (apenas código sem descrição)." if not titulo_masc else "",
    })
    if not titulo_masc:
        add_pendencia("TITULO", codigo, "titulo_normalizado", "Código de título sem descrição no catálogo SITAC", "MÉDIA",
                      "Confirmar descrição do título junto ao SITAC/CONFEA.")
    if len(ocorrencias) > 1:
        add_pendencia("TITULO", codigo, "codigo_titulo", f"Código de título com {len(ocorrencias)} ocorrências no catálogo de origem", "BAIXA",
                      "Descrição principal mantida em TITULOS_PROFISSIONAIS; demais ocorrências preservadas em TITULOS_EQUIVALENCIAS.")
    for extra in extras + (vazias if preenchidas else []):
        if extra is principal:
            continue
        TITULOS_EQUIV_ROWS.append({
            "codigo_titulo": codigo,
            "titulo_alternativo": extra["titulo"] or "(sem descrição — apenas código)",
            "tratamento_alternativo": extra["tratamento"],
            "nivel_alternativo": extra["nivel"],
            "fonte_origem": "FONTE-04",
            "observacao": "Ocorrência adicional do mesmo código no catálogo de origem (duplicidade preservada para auditoria).",
        })

print("TITULOS_PROFISSIONAIS: %d códigos | TITULOS_EQUIVALENCIAS: %d linhas" % (len(TITULOS_ROWS), len(TITULOS_EQUIV_ROWS)))
TITULOS_BY_CODIGO = {t["codigo_titulo"]: t for t in TITULOS_ROWS}

# ---------------------------------------------------------------------------
# 5) CURSOS + CURSO_TITULO  (fonte: Consolidado_EXEMPLO.csv — 7 registros de exemplo)
#    Regra: NUNCA tratados como cadastro validado. Vinculados ao campus real via
#    CNPJ + município (chave confiavel presente nas duas fontes).
# ---------------------------------------------------------------------------
CAMPI_BY_IES = {}
for c in CAMPI_ROWS:
    CAMPI_BY_IES.setdefault(c["id_ies"], []).append(c)

with open(os.path.join(FONTE_DIR, "Consolidado_EXEMPLO.csv"), encoding="utf-8", newline="") as f:
    exemplo_rows = list(csv.DictReader(f))

for i, row in enumerate(exemplo_rows, start=1):
    cnpj_norm = cnpj_digits(row["CNPJ"])
    ies_row = CNPJ_TO_IES.get(cnpj_norm)
    id_curso = stable_id("CURSO", "exemplo", str(i))

    id_campus_ref = ""
    municipio_exemplo = norm_ws(row["CAMPUS_CIDADE"])
    if ies_row:
        candidatos = CAMPI_BY_IES.get(ies_row["id_ies"], [])
        casado = next((c for c in candidatos if strip_accents_upper(c["municipio"]) == strip_accents_upper(municipio_exemplo)), None)
        if not casado and candidatos:
            casado = candidatos[0]
        if casado:
            id_campus_ref = casado["id_campus"]

    nivel = norm_ws(row["NIVEL_CURSO"]).upper() or PENDENTE
    modalidade_raw = strip_accents_upper(row["MODALIDADE"])
    if "EAD" in modalidade_raw or "DISTANCIA" in modalidade_raw:
        modalidade = "EAD"
    elif "SEMI" in modalidade_raw:
        modalidade = "SEMIPRESENCIAL"
    elif "PRESENCIAL" in modalidade_raw:
        modalidade = "PRESENCIAL"
    else:
        modalidade = "NÃO INFORMADA"

    CURSOS_ROWS.append({
        "id_curso": id_curso,
        "id_campus": id_campus_ref or PENDENTE,
        "nome_curso_original": row["CURSO"],
        "nome_curso_normalizado": norm_ws(row["CURSO"]).upper(),
        "grau_academico": nivel,
        "nivel_curso": nivel,
        "modalidade_oferta": modalidade,
        "situacao_curso": PENDENTE,
        "codigo_curso_crea": "",
        "codigo_mec_curso": "",
        "data_inicio": "",
        "data_fim": "",
        "ato_cadastral": PENDENTE,
        "numero_processo": PENDENTE,
        "fonte_origem": "FONTE-05",
        "data_fonte": "2025-09-09",
        "status_validacao": PENDENTE,
        "publicar_dashboard": "NÃO",
    })
    if not id_campus_ref:
        add_pendencia("CURSO", id_curso, "id_campus", "Não foi possível localizar automaticamente o campus correspondente na base de instituições", "MÉDIA",
                      "Confirmar manualmente instituição/campus deste curso de exemplo.")
    else:
        add_pendencia("CURSO", id_curso, "situacao_curso", "Curso de exemplo (fonte: Consolidado/EXEMPLO) — não é cadastro validado do SITAC", "ALTA",
                      "Extrair do SITAC um relatório de cursos oficial (instituição, campus, curso, situação, modalidade, título, ato) antes de publicar.")

    # ---- CURSO_TITULO ----
    codigo_original = re.sub(r"\D", "", row["CODIGO_TITULO"])
    titulo_ref = TITULOS_BY_CODIGO.get(codigo_original)
    id_ct = stable_id("CT", id_curso, codigo_original)
    if titulo_ref:
        CURSO_TITULO_ROWS.append({
            "id_curso_titulo": id_ct,
            "id_curso": id_curso,
            "codigo_titulo": codigo_original,
            "tipo_vinculo": "INFORMADO NA FONTE DE EXEMPLO",
            "fonte_vinculo": "FONTE-05",
            "data_validacao": "",
            "status_validacao": PENDENTE,
            "observacao": "Vínculo curso→título extraído do exemplo Consolidado; não confirmado no SITAC.",
        })
    else:
        add_pendencia("CURSO_TITULO", id_ct, "codigo_titulo", f"Código de título '{codigo_original}' do exemplo não existe no catálogo TITULOS_PROFISSIONAIS", "ALTA",
                      "Confirmar código correto no catálogo SITAC.")

# Divergência conhecida: exemplo do IFBA usa 1211102 (Eng. Industrial Eletrônica)
# para o curso "ENGENHARIA INDUSTRIAL - ELÉTRICA", quando o catálogo associa
# "Elétrica" ao código 1211101. Registrada como pendência, SEM sobrescrever
# silenciosamente o dado de origem.
for row in CURSO_TITULO_ROWS:
    curso = next(c for c in CURSOS_ROWS if c["id_curso"] == row["id_curso"])
    if "ELÉTRICA" in strip_accents_upper(curso["nome_curso_original"]) and row["codigo_titulo"] == "1211102":
        add_pendencia("CURSO_TITULO", row["id_curso_titulo"], "codigo_titulo",
                      "Possível divergência: curso menciona modalidade 'Elétrica' mas está vinculado ao código 1211102 (Engenheiro Industrial Eletrônica); o catálogo associa 'Elétrica' ao código 1211101 (Engenheiro Industrial - Elétrica)",
                      "ALTA", "Confirmar junto ao CREA-BA/SITAC qual código está correto antes de publicar.",
                      observacao="Código sugerido para verificação: 1211101.")

print("CURSOS: %d (todos de exemplo, não publicáveis) | CURSO_TITULO: %d" % (len(CURSOS_ROWS), len(CURSO_TITULO_ROWS)))

# ---------------------------------------------------------------------------
# 6) BASE_LEGAL + ATRIBUICOES + TITULO_ATRIBUICAO_LEGAL
#    Apenas normas reais, confirmadas via pesquisa em fonte oficial nesta sessão.
#    Nenhum texto de "lei tal"/"art. X e Y" da fonte de exemplo foi usado.
# ---------------------------------------------------------------------------
BASE_LEGAL_ROWS.extend([
    {
        "id_norma": "NORMA-01", "tipo_norma": "LEI", "numero_norma": "5.194", "ano": "1966",
        "orgao_emissor": "Congresso Nacional",
        "ementa": "Regula o exercício das profissões de Engenheiro, Arquiteto e Engenheiro-Agrônomo, e dá outras providências.",
        "url_fonte_oficial": "https://www.planalto.gov.br/ccivil_03/leis/l5194.htm",
        "situacao_norma": "VIGENTE", "data_consulta": RUN_DATE,
        "observacao": "Base legal geral do exercício profissional (art. 7º trata das atividades privativas/atribuições).",
    },
    {
        "id_norma": "NORMA-02", "tipo_norma": "RESOLUÇÃO CONFEA", "numero_norma": "218", "ano": "1973",
        "orgao_emissor": "CONFEA",
        "ementa": "Discrimina atividades das diferentes modalidades profissionais da Engenharia, Arquitetura e Agronomia.",
        "url_fonte_oficial": "https://normativos.confea.org.br/Ementas/Visualizar?id=266",
        "situacao_norma": "PARCIALMENTE REVOGADA (arts. 8º e 9º revogados pela Resolução CONFEA nº 1.156/2025; art. 24 revogado pela Resolução 1.057/2014)",
        "data_consulta": RUN_DATE,
        "observacao": "Norma histórica citada nos arquivos de origem como \"lei/artigo tal\" (texto genérico); os artigos 8º/9º relativos a Engenheiro Eletricista/Eletrônico não estão mais em vigor.",
    },
    {
        "id_norma": "NORMA-03", "tipo_norma": "RESOLUÇÃO CONFEA", "numero_norma": "1.156", "ano": "2025",
        "orgao_emissor": "CONFEA",
        "ementa": "Consolida normativos acerca das atividades e competências profissionais dos engenheiros da modalidade eletricista (eletricista, eletrônica, computação, automação, energia, software, biomédica e industrial elétrica).",
        "url_fonte_oficial": "https://normativos.confea.org.br/Ementas/Visualizar?id=82360",
        "situacao_norma": "VIGENTE", "data_consulta": RUN_DATE,
        "observacao": "Norma vigente que substitui os arts. 8º/9º da Resolução 218/1973 para os títulos de engenharia elétrica presentes nesta base. Publicada no DOU em 31/10/2025.",
    },
])

ATRIBUICOES_ROWS.extend([
    {"id_atribuicao": "ATRIB-01", "descricao_atribuicao": "Geração, transmissão, distribuição e utilização de energia elétrica; equipamentos, materiais e máquinas elétricas; sistemas de medição e controle elétricos.", "area_atuacao": "Engenharia Elétrica — Eletrotécnica", "observacao": "Correspondia ao art. 8º da Resolução 218/1973 (revogado); atualmente tratada pela Resolução CONFEA nº 1.156/2025."},
    {"id_atribuicao": "ATRIB-02", "descricao_atribuicao": "Materiais e equipamentos eletrônicos; sistemas de comunicação e telecomunicações; sistemas de medição e controle elétrico e eletrônico.", "area_atuacao": "Engenharia Elétrica — Eletrônica/Telecomunicações", "observacao": "Correspondia ao art. 9º da Resolução 218/1973 (revogado); atualmente tratada pela Resolução CONFEA nº 1.156/2025."},
    {"id_atribuicao": "ATRIB-03", "descricao_atribuicao": "Atividades de engenharia de computação (hardware, software e sistemas computacionais).", "area_atuacao": "Engenharia de Computação", "observacao": "Atribuições tornadas definitivas pela Resolução CONFEA nº 1.156/2025, conforme nota oficial do CONFEA."},
])

# Vínculo título -> atribuição -> norma. A norma e a existência do vínculo estão
# confirmadas por fonte oficial; o artigo específico da Resolução 1.156/2025 para
# cada título ainda depende de leitura integral do anexo da norma — por isso o
# campo "artigo" permanece PENDENTE em vez de ser inferido.
VINCULOS_TITULO_ATRIBUICAO = [
    ("1210800", "ATRIB-01", "NORMA-03"),  # Engenheiro Eletricista
    ("1210802", "ATRIB-01", "NORMA-03"),  # Engenheiro Eletricista - Eletrotécnico
    ("1211101", "ATRIB-01", "NORMA-03"),  # Engenheiro Industrial - Elétrica
    ("1210801", "ATRIB-02", "NORMA-03"),  # Engenheiro Eletricista Eletrônica
    ("1211102", "ATRIB-02", "NORMA-03"),  # Engenheiro Industrial Eletrônica
    ("1210900", "ATRIB-02", "NORMA-03"),  # Engenheiro em Eletrônica
    ("1210100", "ATRIB-03", "NORMA-03"),  # Engenheiro de Computação
]
for i, (codigo_titulo, id_atrib, id_norma) in enumerate(VINCULOS_TITULO_ATRIBUICAO, start=1):
    if codigo_titulo not in TITULOS_BY_CODIGO:
        continue
    id_vinculo = f"TAL-{i:03d}"
    TITULO_ATRIB_LEGAL_ROWS.append({
        "id_vinculo_legal": id_vinculo,
        "codigo_titulo": codigo_titulo,
        "id_atribuicao": id_atrib,
        "id_norma": id_norma,
        "artigo": PENDENTE,
        "paragrafo": "", "inciso": "", "alinea": "",
        "texto_resumido": next(a["descricao_atribuicao"] for a in ATRIBUICOES_ROWS if a["id_atribuicao"] == id_atrib),
        "url_fonte_oficial": next(n["url_fonte_oficial"] for n in BASE_LEGAL_ROWS if n["id_norma"] == id_norma),
        "status_validacao": "NORMA CONFIRMADA; ARTIGO ESPECÍFICO PENDENTE",
        "data_validacao": RUN_DATE,
        "responsavel_validacao": "Pesquisa assistida (fonte oficial CONFEA) — requer validação humana do artigo exato.",
        "observacao": "",
    })
    add_pendencia("TITULO_ATRIBUICAO_LEGAL", id_vinculo, "artigo",
                  "Norma confirmada oficialmente, mas o artigo/parágrafo específico da Resolução CONFEA 1.156/2025 para este título ainda não foi extraído", "MÉDIA",
                  "Ler o texto integral da Resolução 1.156/2025 (ou seu anexo) e preencher o campo 'artigo'.")

print("BASE_LEGAL: %d normas | ATRIBUICOES: %d | TITULO_ATRIBUICAO_LEGAL: %d" % (
    len(BASE_LEGAL_ROWS), len(ATRIBUICOES_ROWS), len(TITULO_ATRIB_LEGAL_ROWS)))

# ---------------------------------------------------------------------------
# 7) Pendências adicionais (campus sem coordenadas, cursos sem base legal, etc.)
# ---------------------------------------------------------------------------
for c in CAMPI_ROWS:
    if c["situacao_campus"] not in ("TRANSFERIDO",) and not c["latitude"] and c["municipio"]:
        add_pendencia("CAMPUS", c["id_campus"], "latitude/longitude", "Município identificado mas sem coordenada associada", "BAIXA",
                      "Investigar falha de correspondência no lookup de municípios IBGE.")
    if c["situacao_campus"] not in ("TRANSFERIDO",) and not c["municipio"]:
        c["publicar_dashboard"] = "NÃO"

for curso in CURSOS_ROWS:
    tem_titulo = any(ct["id_curso"] == curso["id_curso"] for ct in CURSO_TITULO_ROWS)
    if not tem_titulo:
        add_pendencia("CURSO", curso["id_curso"], "codigo_titulo", "Curso sem título profissional vinculado", "ALTA",
                      "Confirmar título/código no SITAC.")
    tem_base_legal = any(
        ct["id_curso"] == curso["id_curso"] and
        any(tal["codigo_titulo"] == ct["codigo_titulo"] for tal in TITULO_ATRIB_LEGAL_ROWS)
        for ct in CURSO_TITULO_ROWS
    )
    if not tem_base_legal:
        add_pendencia("CURSO", curso["id_curso"], "base_legal", "Curso sem base legal (norma/artigo) confirmada e vinculada", "ALTA",
                      "Vincular a BASE_LEGAL/TITULO_ATRIBUICAO_LEGAL somente após confirmação oficial.")

for ies in IES_ROWS:
    campi_da_ies = CAMPI_BY_IES.get(ies["id_ies"], [])
    if not campi_da_ies:
        add_pendencia("IES", ies["id_ies"], "campi", "Instituição sem nenhum campus associado", "MÉDIA", "Revisar linha de origem.")

print("PENDENCIAS: %d registros" % len(PENDENCIAS_ROWS))

# ---------------------------------------------------------------------------
# 8) BASE_DASHBOARD — tabela desnormalizada para consumo do dashboard
# ---------------------------------------------------------------------------
CT_BY_CURSO = {}
for ct in CURSO_TITULO_ROWS:
    CT_BY_CURSO.setdefault(ct["id_curso"], []).append(ct)

TAL_BY_TITULO = {}
for tal in TITULO_ATRIB_LEGAL_ROWS:
    TAL_BY_TITULO.setdefault(tal["codigo_titulo"], []).append(tal)

CURSOS_BY_CAMPUS = {}
for curso in CURSOS_ROWS:
    CURSOS_BY_CAMPUS.setdefault(curso["id_campus"], []).append(curso)

dash_seq = 0


def novo_id_dashboard():
    global dash_seq
    dash_seq += 1
    return f"PUB-{dash_seq:05d}"


def campus_publicavel(campus):
    return campus["situacao_campus"] == "SITUAÇÃO NÃO INFORMADA" and bool(campus["municipio"])


for ies in IES_ROWS:
    campi = CAMPI_BY_IES.get(ies["id_ies"], [])
    if not campi:
        campi = [None]
    for campus in campi:
        cursos_do_campus = CURSOS_BY_CAMPUS.get(campus["id_campus"], []) if campus else []
        linhas_curso = cursos_do_campus if cursos_do_campus else [None]
        for curso in linhas_curso:
            titulos_do_curso = CT_BY_CURSO.get(curso["id_curso"], []) if curso else []
            linhas_titulo = titulos_do_curso if titulos_do_curso else [None]
            for ct in linhas_titulo:
                titulo_ref = TITULOS_BY_CODIGO.get(ct["codigo_titulo"]) if ct else None
                tal_ref = TAL_BY_TITULO.get(ct["codigo_titulo"], [None])[0] if ct else None
                norma_ref = None
                if tal_ref:
                    norma_ref = next((n for n in BASE_LEGAL_ROWS if n["id_norma"] == tal_ref["id_norma"]), None)

                publicavel = "SIM" if (
                    ies["publicar_dashboard"] == "SIM"
                    and campus is not None and campus_publicavel(campus)
                    and (curso is None or curso["publicar_dashboard"] == "SIM")
                ) else "NÃO"

                status_cadastral = "PUBLICÁVEL" if publicavel == "SIM" else (
                    "TRANSFERIDO PARA OUTRO CONSELHO" if campus and campus["situacao_campus"] == "TRANSFERIDO" else
                    "REVISAR ANTES DE PUBLICAR"
                )
                status_legal = tal_ref["status_validacao"] if tal_ref else PENDENTE

                BASE_DASHBOARD_ROW = {
                    "id_registro_dashboard": novo_id_dashboard(),
                    "id_ies": ies["id_ies"],
                    "nome_ies": ies["nome_normalizado"],
                    "sigla_ies": ies["sigla"],
                    "categoria_academica": ies["categoria_academica"],
                    "natureza_administrativa": ies["natureza_administrativa"],
                    "cnpj": ies["cnpj_normalizado"],
                    "codigo_mec_ies": ies["codigo_mec"],
                    "id_campus": campus["id_campus"] if campus else "",
                    "nome_campus": campus["campus_normalizado"] if campus else "",
                    "municipio": campus["municipio"] if campus else "",
                    "uf": campus["uf"] if campus else "",
                    "endereco": campus["endereco"] if campus else "",
                    "latitude": campus["latitude"] if campus else "",
                    "longitude": campus["longitude"] if campus else "",
                    "id_curso": curso["id_curso"] if curso else "",
                    "nome_curso": curso["nome_curso_normalizado"] if curso else "",
                    "grau_academico": curso["grau_academico"] if curso else "",
                    "nivel_curso": curso["nivel_curso"] if curso else "",
                    "modalidade_oferta": curso["modalidade_oferta"] if curso else "",
                    "situacao_curso": curso["situacao_curso"] if curso else "",
                    "codigo_titulo": ct["codigo_titulo"] if ct else "",
                    "titulo_profissional": titulo_ref["titulo_normalizado"] if titulo_ref else "",
                    "norma": (f"{norma_ref['tipo_norma']} {norma_ref['numero_norma']}/{norma_ref['ano']}" if norma_ref else ""),
                    "artigo": tal_ref["artigo"] if tal_ref else "",
                    "atribuicao_resumida": tal_ref["texto_resumido"] if tal_ref else "",
                    "url_fonte_legal": tal_ref["url_fonte_oficial"] if tal_ref else "",
                    "status_validacao_cadastral": status_cadastral,
                    "status_validacao_legal": status_legal,
                    "data_atualizacao": RUN_DATE,
                    "publicar_dashboard": publicavel,
                }
                globals().setdefault("BASE_DASHBOARD_ROWS", []).append(BASE_DASHBOARD_ROW)

BASE_DASHBOARD_ROWS = globals()["BASE_DASHBOARD_ROWS"]
print("BASE_DASHBOARD: %d linhas" % len(BASE_DASHBOARD_ROWS))

# ---------------------------------------------------------------------------
# 9) Gravação da planilha
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def escreve_aba(wb, nome_aba, linhas, colunas=None, primeira=False):
    ws = wb.active if primeira else wb.create_sheet(title=nome_aba)
    if primeira:
        ws.title = nome_aba
    if not linhas:
        if colunas:
            ws.append(colunas)
            for i, _ in enumerate(colunas, start=1):
                ws.cell(row=1, column=i).fill = HEADER_FILL
                ws.cell(row=1, column=i).font = HEADER_FONT
        return ws
    cols = colunas or [k for k in linhas[0].keys() if not k.startswith("_")]
    ws.append(cols)
    for i, _ in enumerate(cols, start=1):
        ws.cell(row=1, column=i).fill = HEADER_FILL
        ws.cell(row=1, column=i).font = HEADER_FONT
    for linha in linhas:
        ws.append([str(linha.get(c, "")) if linha.get(c, "") != "" else "" for c in cols])
    for i, col in enumerate(cols, start=1):
        largura = min(60, max(12, len(col) + 2, *(len(str(l.get(col, ""))) for l in linhas[:200])))
        ws.column_dimensions[get_column_letter(i)].width = largura
    ws.freeze_panes = "A2"
    return ws


os.makedirs(PLANILHA_DIR, exist_ok=True)
wb = Workbook()

escreve_aba(wb, "IES", IES_ROWS, primeira=True)
escreve_aba(wb, "CAMPI", CAMPI_ROWS)
escreve_aba(wb, "CURSOS", CURSOS_ROWS)
escreve_aba(wb, "TITULOS_PROFISSIONAIS", TITULOS_ROWS)
escreve_aba(wb, "TITULOS_EQUIVALENCIAS", TITULOS_EQUIV_ROWS,
            colunas=["codigo_titulo", "titulo_alternativo", "tratamento_alternativo", "nivel_alternativo", "fonte_origem", "observacao"])
escreve_aba(wb, "CURSO_TITULO", CURSO_TITULO_ROWS)
escreve_aba(wb, "BASE_LEGAL", BASE_LEGAL_ROWS)
escreve_aba(wb, "ATRIBUICOES", ATRIBUICOES_ROWS)
escreve_aba(wb, "TITULO_ATRIBUICAO_LEGAL", TITULO_ATRIB_LEGAL_ROWS)
escreve_aba(wb, "FONTES", FONTES_ROWS)
escreve_aba(wb, "PENDENCIAS", PENDENCIAS_ROWS)
escreve_aba(wb, "BASE_DASHBOARD", BASE_DASHBOARD_ROWS)

# aba RESUMO no início
ws_resumo = wb.create_sheet(title="RESUMO", index=0)
ws_resumo.append(["BASE ÚNICA PADRONIZADA — INSTITUIÇÕES, CURSOS, TÍTULOS E ATRIBUIÇÕES DO CREA-BA"])
ws_resumo["A1"].font = Font(bold=True, size=14)
ws_resumo.append([f"Gerado em {RUN_TIMESTAMP} a partir dos arquivos em dados/fonte/ (ver aba FONTES)."])
ws_resumo.append([])
resumo_pares = [
    ("Instituições (IES) únicas", len(IES_ROWS)),
    ("Campi únicos", len(CAMPI_ROWS)),
    ("Municípios identificados", len({c["municipio"] for c in CAMPI_ROWS if c["municipio"]})),
    ("Linhas em BASE_DASHBOARD", len(BASE_DASHBOARD_ROWS)),
    ("Registros publicáveis (publicar_dashboard = SIM)", sum(1 for r in BASE_DASHBOARD_ROWS if r["publicar_dashboard"] == "SIM")),
    ("Cursos de exemplo (não publicáveis)", len(CURSOS_ROWS)),
    ("Códigos de título profissional", len(TITULOS_ROWS)),
    ("Normas confirmadas em BASE_LEGAL", len(BASE_LEGAL_ROWS)),
    ("Pendências abertas", len(PENDENCIAS_ROWS)),
]
ws_resumo.append(["Indicador", "Valor"])
ws_resumo["A4"].font = HEADER_FONT
ws_resumo["A4"].fill = HEADER_FILL
ws_resumo["B4"].font = HEADER_FONT
ws_resumo["B4"].fill = HEADER_FILL
for label, val in resumo_pares:
    ws_resumo.append([label, val])
ws_resumo.column_dimensions["A"].width = 48
ws_resumo.column_dimensions["B"].width = 14

wb.save(OUT_XLSX)
print("\nPlanilha gravada em: %s" % OUT_XLSX)
